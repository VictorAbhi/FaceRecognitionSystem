from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os
import cv2  # Ensure cv2 is imported for video capture handling


class AttendanceMarker:
    def __init__(self, root, excel_file="attendance.xlsx"):
        # Initialize the Tkinter root and the excel file path
        self.root = root
        self.excel_file = excel_file
        # Check if the excel file exists, create it if not
        if not os.path.exists(self.excel_file):
            self.create_excel_file()

    def create_excel_file(self):
        # Create a new Excel workbook and worksheet with headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["ID", "Name", "Date", "Time"])
        workbook.save(self.excel_file)

    def mark_attendance(self, i, n, videoCap):
        workbook = openpyxl.load_workbook(self.excel_file)
        sheet = workbook.active
        current_date = datetime.now().strftime("%d/%m/%Y")

        # Check if the attendance has already been marked today for this ID
        for row in sheet.iter_rows(values_only=True):
            if row[0] == i and row[2] == current_date:
                # Attendance already marked, ask if they want to proceed to the next attendee
                proceed = messagebox.askyesno(
                    title=f"Attendance already marked for {n} (ID: {i})",
                    message="Attendance is already marked for today. Do you want to proceed to the next attendee?"
                )
                if proceed:
                    return
                else:
                    # If No, release the camera and close OpenCV windows
                    videoCap.release()  # Release the video capture
                    cv2.destroyAllWindows()  # Close all OpenCV windows
                    self.root.quit()  # Close the Tkinter window (terminate the program)
                    return

        # If attendance has not been marked for today, proceed to mark it
        current_time = datetime.now().strftime("%H:%M:%S")
        sheet.append([i, n, current_date, current_time])
        workbook.save(self.excel_file)

        sheet.append([])  # This adds a blank row
        total_attendees_today = sum(1 for row in sheet.iter_rows(values_only=True) if row[2] == current_date)
        sheet.append([f"Day Summary for {current_date}", f"Total Attendees: {total_attendees_today}"])

        # Show a confirmation dialog after marking attendance
        messagebox.showinfo(
            title=f"Attendance marked for {n} (ID: {i})",
            message=f"Attendance marked for {n} (ID: {i}) on {current_date} at {current_time}. Do you want to proceed to the next attendee?"
        )
